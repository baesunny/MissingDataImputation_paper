"""
고비용 공정 데이터의 결측치 대체 통합 스크립트.
MICE(sklearn), missForest(miceforest), 1D-CNN·MLP(PyTorch) 중 하나를 선택해
독립변수 결측을 대체하고 Excel·PDF 리포트를 생성한다.
"""
import os
import sys
import glob

# -----------------------------------------------------------------------------
# MacOS Environment Fix (Must be before other imports)
# -----------------------------------------------------------------------------
if sys.platform == 'darwin':
    for env_var in ['DYLD_LIBRARY_PATH', 'DYLD_FALLBACK_LIBRARY_PATH']:
        if env_var in os.environ:
            del os.environ[env_var]

import pandas as pd
import numpy as np
import json
import re

import matplotlib.pyplot as plt
import seaborn as sns
from matplotlib.backends.backend_pdf import PdfPages
from openpyxl import load_workbook
from openpyxl.styles import Font

# ML/DL Libraries
try:
    import miceforest as mf
    from sklearn.experimental import enable_iterative_imputer
    from sklearn.impute import IterativeImputer
    from sklearn.preprocessing import StandardScaler
    import torch
    import torch.nn as nn
    import torch.optim as optim
    from torch.utils.data import TensorDataset, DataLoader
except ImportError as e:
    print(f"[Critical] 필수 라이브러리가 없다: {e}")
    sys.exit(1)

# -----------------------------------------------------------------------------
# Global Variables (Configuration)
# -----------------------------------------------------------------------------
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SOURCE_DIR = os.path.join(BASE_DIR, "sources")
RESULT_DIR = os.path.join(BASE_DIR, "outputs", "imputation_runs")
CONFIG_FILE = os.path.join(BASE_DIR, "config.json")
IMPUTATION_METHODS = ["MICE", "missForest", "1D-CNN", "MLP"]

# Matplotlib 한글 폰트 설정 (macOS 기준, Windows는 Malgun Gothic 등으로 변경 가능)
plt.rcParams['font.family'] = 'NanumGothic'
plt.rcParams['axes.unicode_minus'] = False

# -----------------------------------------------------------------------------
# Helper Functions: Config & Parsing
# -----------------------------------------------------------------------------
def load_config():
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception: pass
    return {}

def save_config(config):
    try:
        with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
            json.dump(config, f, indent=4, ensure_ascii=False)
    except Exception: pass

def parse_column_selection(input_str, total_columns):
    selected_indices = set()
    parts = [p.strip() for p in input_str.split(',')]
    for part in parts:
        if not part: continue
        if '-' in part:
            try:
                start, end = map(int, part.split('-'))
                selected_indices.update(range(start - 1, end))
            except ValueError: pass
        else:
            try:
                idx = int(part)
                selected_indices.add(idx - 1)
            except ValueError: pass
    return sorted([i for i in selected_indices if 0 <= i < total_columns])

# -----------------------------------------------------------------------------
# Imputation Models (PyTorch)
# -----------------------------------------------------------------------------
class Simple1DCNN(nn.Module):
    def __init__(self, input_dim):
        super(Simple1DCNN, self).__init__()
        self.conv1 = nn.Conv1d(input_dim, 64, kernel_size=3, padding=1)
        self.relu = nn.ReLU()
        self.pool = nn.MaxPool1d(2) # Output ~ len/2
        # After pool, dimension depends... but here input is (Batch, Feat, 1) usually?
        # Tabular 데이터에서는 1D CNN이 feature를 채널 또는 시퀀스로 취급한다.
        # 여기서는 (batch, 1, features) 형태로 단순화한다.
        # Refactoring Concept: Input (Batch, 1, Features) -> Conv -> (Batch, 64, Features)
        
        # Simplified Logic matching reference R code spirit but adapted:
        # Ref Code: Conv1d(in=Feat, out=64, k=3) -> Pool -> ...
        # If we treat input as (Batch, Features), unsqueeze to (Batch, Features, 1) fails kernel 3 if features < 3.
        # Let's treat Input as (Batch, 1, Features).
        self.conv_layer = nn.Sequential(
            nn.Conv1d(1, 16, kernel_size=3, padding=1),
            nn.ReLU(),
            nn.Flatten()
        )
        self.fc = nn.Sequential(
            nn.Linear(16 * input_dim, 32), # Rough estimate, dynamically handled in forward preference usually
            nn.ReLU(),
            nn.Linear(32, input_dim) # Predicting reconstruction
        )
        self.input_dim = input_dim

    def forward(self, x):
        # x: (Batch, Features) -> (Batch, 1, Features)
        x_in = x.unsqueeze(1)
        x_conv = self.conv_layer(x_in)
        out = self.fc(x_conv)
        return out

class SimpleMLP(nn.Module):
    def __init__(self, input_dim):
        super(SimpleMLP, self).__init__()
        self.net = nn.Sequential(
            nn.Linear(input_dim, 64),
            nn.BatchNorm1d(64),
            nn.ReLU(),
            nn.Dropout(0.2),
            nn.Linear(64, 32),
            nn.ReLU(),
            nn.Linear(32, input_dim)
        )
    def forward(self, x):
        return self.net(x)

def train_dl_model(model, data_loader, epochs=100, device='cpu'):
    model.to(device)
    optimizer = optim.Adam(model.parameters(), lr=0.001)
    criterion = nn.MSELoss()
    model.train()
    
    for epoch in range(epochs):
        total_loss = 0
        for batch_x in data_loader:
            batch_x = batch_x[0].to(device)
            optimizer.zero_grad()
            output = model(batch_x)
            loss = criterion(output, batch_x) # Autoencoder Style Reconstruction
            loss.backward()
            optimizer.step()
            total_loss += loss.item()
    return model

def impute_dl(df, model_type='CNN'):
    # Simple strategy: Fill NA with mean -> Train Autoencoder -> Predict -> Replace NA
    data_np = df.values.copy()
    rows, cols = data_np.shape
    
    # Pre-fill NA with mean for training
    scaler = StandardScaler()
    
    # 완전히 비어 있는 열이 있으면 mean 대체 전에 처리한다 (전처리 단계에서 제거하는 것이 원칙).
    full_mask = ~np.isnan(data_np)
    
    # Temp Imputation (Mean)
    temp_filled = data_np.copy()
    col_means = np.nanmean(temp_filled, axis=0)
    inds = np.where(np.isnan(temp_filled))
    temp_filled[inds] = np.take(col_means, inds[1])
    
    # Scale
    scaled_data = scaler.fit_transform(temp_filled)
    tensor_data = torch.FloatTensor(scaled_data)
    
    dataset = TensorDataset(tensor_data)
    loader = DataLoader(dataset, batch_size=32, shuffle=True)
    
    # Init Model
    device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')
    if model_type == '1D-CNN':
        model = Simple1DCNN(cols)
    else: # MLP
        model = SimpleMLP(cols)
        
    # Train
    print(f"  [Training] {model_type} on {device}...")
    model = train_dl_model(model, loader, epochs=50, device=device)
    
    # Predict (Reconstruct)
    model.eval()
    with torch.no_grad():
        reconstructed = model(tensor_data.to(device)).cpu().numpy()
        
    # Inverse Scale
    reconstructed_original = scaler.inverse_transform(reconstructed)
    
    # Replace only NAs
    final_data = data_np.copy()
    final_data[inds] = reconstructed_original[inds]
    
    return pd.DataFrame(final_data, columns=df.columns)

# -----------------------------------------------------------------------------
# Imputation Logic
# -----------------------------------------------------------------------------
def run_imputation(df, method, x_cols, y_col):
    """
    선택한 대체 기법을 독립변수(X) 열에 적용한다.
    종속변수(Y)는 본 스크립트에서 갱신하지 않는다.
    """
    print(f"\n[Process] '{method}' 기법으로 결측치 대체를 수행한다...")
    
    # Filter Data
    target_df = df[x_cols].copy()
    
    # Check if any missing
    if target_df.isnull().sum().sum() == 0:
        print("  [Info] 선택된 독립변수에 결측치가 없다.")
        return target_df

    imputed_df = target_df.copy()
    
    if method == "MICE":
        # sklearn IterativeImputer 사용 (BayesianRidge 계열, MICE와 유사한 연쇄 대체)
        # 숫자형 열만 대상으로 한다.
        numeric_cols = target_df.select_dtypes(include=[np.number]).columns
        if len(numeric_cols) != len(target_df.columns):
            excluded = list(set(target_df.columns) - set(numeric_cols))
            print(f"  [Warning] MICE(sklearn)는 숫자형 열만 지원한다. 제외 열: {excluded}")
            target_df_numeric = target_df[numeric_cols]
        else:
            target_df_numeric = target_df

        imputer = IterativeImputer(random_state=42)
        imputed_vals = imputer.fit_transform(target_df_numeric)
        
        # Update specific columns in the original full dataframe structure
        # imputed_df is already a copy of target_df, so non-numeric cols are preserved.
        imputed_df[numeric_cols] = imputed_vals
        
    elif method == "missForest":
        # miceforest는 범주형 열에 category dtype이 필요하다.
        # object 열을 category로 변환한다.
        obj_cols = target_df.select_dtypes(include=['object']).columns
        if len(obj_cols) > 0:
            print(f"  [Info] object 열을 category로 변환한다: {list(obj_cols)}")
            for col in obj_cols:
                target_df[col] = target_df[col].astype('category')

        # miceforest(LightGBM 기반)로 missForest에 준하는 다변량 대체를 수행한다.
        kds = mf.ImputationKernel(
            target_df,
            random_state=42
        )
        kds.mice(3)
        imputed_df = kds.complete_data()
        
    elif method in ["1D-CNN", "MLP"]:
        # 딥러닝 경로는 숫자형 열만 지원한다.
        numeric_cols = target_df.select_dtypes(include=[np.number]).columns
        if len(numeric_cols) != len(target_df.columns):
            print("  [Warning] DL 방식은 숫자형 열만 지원한다. 비숫자형은 제외 후 진행한다.")
            data_to_dl = target_df[numeric_cols]
        else:
            data_to_dl = target_df
            
        imputed_dl = impute_dl(data_to_dl, model_type=method)
        # Updates only numeric
        imputed_df.update(imputed_dl)
        
    return imputed_df

# -----------------------------------------------------------------------------
# Reporting
# -----------------------------------------------------------------------------
def save_excel_highlighted(original, imputed, filename):
    """
    DataFrame을 Excel로 저장하고, original에서 결측이었던 셀은 빨간색 글자로 표시한다.
    """
    # 1. Save Basic Excel
    imputed.to_excel(filename, index=False)
    
    # 2. Open with OpenPyXML to Apply Style
    wb = load_workbook(filename)
    ws = wb.active
    
    red_font = Font(color="FF0000")
    
    # Iterate over original dataframe to find NaN locations
    # Excel rows: 2 to len(df)+1 (Row 1 is header)
    # Excel cols: 1 to len(columns)
    
    # Pre-calculate NaN masks to avoid repeated iloc/checking
    # We iterate by integer index
    n_rows, n_cols = original.shape
    
    # openpyxl은 1-based 인덱스를 사용한다.
    # imputed는 original과 동일한 shape·열 순서를 가진다고 가정한다.
    
    for r in range(n_rows):
        for c in range(n_cols):
            # Check safely for NaN in original
            val = original.iloc[r, c]
            if pd.isna(val):
                # Apply style to corresponding cell in Excel
                # Row: r (0-based) -> r + 2 (Header + 1-based)
                # Col: c (0-based) -> c + 1 (1-based)
                ws.cell(row=r+2, column=c+1).font = red_font
                
    wb.save(filename)
    print(f"[Result] 엑셀 저장 완료(대체값 붉은색 표시): {filename}")

def create_pdf_report(filename, original, imputed, method, x_cols):
    """
    대체 결과 분포를 시각화해 PDF 리포트로 저장한다.
    """
    # 결측이 있었던 독립변수(x_cols)만 시각화 대상으로 한다.
    missing_cols = [col for col in original.columns if col in x_cols and pd.isna(original[col]).any()]
    
    if not missing_cols:
        print("  [Info] 결측이 있는 독립변수가 없어 PDF 리포트를 생성하지 않는다.")
        return
    
    with PdfPages(filename) as pdf:
        # Title Page
        fig = plt.figure(figsize=(11, 8.5))
        fig.text(0.5, 0.7, f"Imputation Report", 
                 ha='center', va='center', fontsize=24, fontweight='bold')
        fig.text(0.5, 0.6, f"Method: {method}", 
                 ha='center', va='center', fontsize=18)
        fig.text(0.5, 0.5, f"Generated: {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')}", 
                 ha='center', va='center', fontsize=14)
        fig.text(0.5, 0.4, f"Imputed Columns: {len(missing_cols)}", 
                 ha='center', va='center', fontsize=14)
        fig.text(0.5, 0.3, f"(Independent Variables Only)", 
                 ha='center', va='center', fontsize=12, style='italic')
        plt.axis('off')
        pdf.savefig(fig)
        plt.close()
        
        # Distribution plots for each imputed column
        for idx, col in enumerate(missing_cols):
            if col not in imputed.columns:
                continue
            
            fig, axes = plt.subplots(1, 2, figsize=(14, 5))
            
            # Left: KDE Plot (Distribution Comparison)
            try:
                # Filter non-null values for Original
                orig_values = original[col].dropna()
                imp_values = imputed[col].dropna()
                
                if len(orig_values) > 0:
                    sns.kdeplot(orig_values, ax=axes[0], label='Original (Non-Missing)', 
                               color='blue', fill=True, alpha=0.3, linewidth=2)
                if len(imp_values) > 0:
                    sns.kdeplot(imp_values, ax=axes[0], label='After Imputation (All)', 
                               color='red', linestyle='--', alpha=0.6, linewidth=2)
                
                axes[0].set_title(f"Distribution Comparison: {col}", fontsize=12, fontweight='bold')
                axes[0].set_xlabel("Value")
                axes[0].set_ylabel("Density")
                axes[0].legend()
                axes[0].grid(True, alpha=0.3)
            except Exception as e:
                axes[0].text(0.5, 0.5, f"Error plotting distribution:\n{str(e)}", 
                           ha='center', va='center', transform=axes[0].transAxes)
            
            # Right: Box Plot (Outlier Comparison)
            try:
                data_to_plot = []
                labels = []
                
                if len(orig_values) > 0:
                    data_to_plot.append(orig_values)
                    labels.append('Original')
                if len(imp_values) > 0:
                    data_to_plot.append(imp_values)
                    labels.append('Imputed')
                
                if data_to_plot:
                    bp = axes[1].boxplot(data_to_plot, tick_labels=labels, patch_artist=True,
                                        boxprops=dict(facecolor='lightblue', alpha=0.6),
                                        medianprops=dict(color='red', linewidth=2))
                    axes[1].set_title(f"Box Plot: {col}", fontsize=12, fontweight='bold')
                    axes[1].set_ylabel("Value")
                    axes[1].grid(True, alpha=0.3, axis='y')
            except Exception as e:
                axes[1].text(0.5, 0.5, f"Error plotting boxplot:\n{str(e)}", 
                           ha='center', va='center', transform=axes[1].transAxes)
            
            # Add missing count info
            missing_count = pd.isna(original[col]).sum()
            total_count = len(original)
            fig.text(0.5, 0.02, 
                    f"Missing Values: {missing_count} / {total_count} ({missing_count/total_count*100:.1f}%)",
                    ha='center', fontsize=10, style='italic')
            
            plt.tight_layout(rect=[0, 0.03, 1, 1])
            pdf.savefig(fig)
            plt.close()
        
        # Summary page
        fig = plt.figure(figsize=(11, 8.5))
        summary_text = f"Summary\n\n"
        summary_text += f"Total Columns Imputed: {len(missing_cols)}\n\n"
        summary_text += "Column List:\n"
        for i, col in enumerate(missing_cols, 1):
            missing_count = pd.isna(original[col]).sum()
            summary_text += f"  {i}. {col}: {missing_count} missing values\n"
        
        fig.text(0.1, 0.9, summary_text, 
                ha='left', va='top', fontsize=11, family='monospace')
        plt.axis('off')
        pdf.savefig(fig)
        plt.close()
        
    print(f"[Report] PDF 리포트 저장 완료: {filename}")

# -----------------------------------------------------------------------------
# Core Features (Updated)
# -----------------------------------------------------------------------------
def setup_environment():
    if not os.path.exists(SOURCE_DIR): pass 
    if not os.path.exists(RESULT_DIR): os.makedirs(RESULT_DIR, exist_ok=True)

def list_excel_files():
    files = []
    if os.path.exists(SOURCE_DIR):
        for ext in ['*.xlsx', '*.xls']:
            files.extend(glob.glob(os.path.join(SOURCE_DIR, ext)))
    return sorted([os.path.basename(f) for f in files])

def select_file():
    files = list_excel_files()
    if not files: return None
    print("\n[Step 1] 분석할 파일을 선택해 주세요:")
    for idx, f in enumerate(files): print(f"{idx + 1}. {f}")
    while True:
        try:
            sel = int(input("\n번호 (종료: q): "))
            if 1 <= sel <= len(files): return os.path.join(SOURCE_DIR, files[sel-1])
        except: sys.exit(0)

def print_columns(df):
    cols = df.columns.tolist()
    print("\n[현재 컬럼 목록]")
    for i in range(0, len(cols), 5):
        chunk = cols[i:i+5]
        line = " | ".join([f"{j+1}. {col}" for j, col in enumerate(chunk, start=i)])
        print(line)

def get_start_row(config):
    s_row = config.get('start_row', 1)
    inp = input(f"\n[Step 2] 데이터 시작 행 [기본: {s_row}]: ").strip()
    return int(inp) if inp else s_row

def get_drop_columns(df, config):
    print_columns(df)
    d_cols = config.get('drop_columns_str', "")
    inp = input(f"\n[Step 3] 제외할 열 (예: 1-3). [이전: '{d_cols}']: ").strip()
    return inp if inp else d_cols

def filter_by_class(df, config):
    """
    Class1·Class2 열이 있으면 고유값을 보여 주고 필터링 여부를 입력받는다.
    """
    for col in ['Class1', 'Class2']:
        if col in df.columns:
            unique_vals = sorted(df[col].dropna().unique().tolist())
            print(f"\n[Step 4] '{col}' 데이터 필터링")
            print(f"고유값 목록: {unique_vals}")
            print("1. 전체 데이터 사용 (필터링 없음)")
            print("2. 특정 값만 선택하여 나머지 제외")
            
            # Load default
            def_sel = config.get(f'filter_{col}_sel', '1')
            def_val = config.get(f'filter_{col}_val', '')

            try:
                sel = input(f"선택 (1/2) [기본: {def_sel}]: ").strip()
                if not sel: sel = def_sel
                
                # Update Config
                config[f'filter_{col}_sel'] = sel
                
                if sel == '2':
                    prompt_val = f"값 [기본: {def_val}]: " if def_val else "값: "
                    print("사용할 값을 입력하세요 (위 목록 중 하나 정확히 입력):")
                    val_str = input(prompt_val).strip()
                    if not val_str and def_val: val_str = def_val
                    
                    # Update Config
                    config[f'filter_{col}_val'] = val_str

                    # Try to match type (int/float/str)
                    target_val = val_str
                    if pd.api.types.is_numeric_dtype(df[col]):
                        try:
                            if '.' in val_str:
                                target_val = float(val_str)
                            else:
                                target_val = int(val_str)
                        except: pass
                    
                    if target_val in unique_vals:
                        initial_len = len(df)
                        df = df[df[col] == target_val].copy()
                        print(f"  Result: {initial_len}행 -> {len(df)}행 (Test 제외)")
                        # Drop the column as it's now constant
                        df.drop(columns=[col], inplace=True)
                        # Reset index for miceforest stability
                        df.reset_index(drop=True, inplace=True)
                        print(f"  Info: '{col}' 열은 독립변수에서 제외되었다.")
                    else:
                        print(f"  Warning: '{target_val}' 값을 찾을 수 없어 전체 데이터를 사용한다.")
                else:
                    print(f"  Info: '{col}' 전체 데이터를 사용한다.")
            except Exception as e:
                print(f"  Error: {e}. 전체 데이터를 사용한다.")
    return df

def get_index_selection(df, config):
    print("\n[Step 4-1] Index(식별자) 변수 지정")
    print_columns(df)
    
    idx_str = config.get('index_vars_str', "")
    inp = input(f"- Index 변수 (분석 제외, 결과 포함) [이전: '{idx_str}']: ").strip()
    idx_str = inp if inp else idx_str
    
    return idx_str

def get_var_selection(df, config):
    print("\n[Step 4-2] 변수 지정 (필터링 후)")
    # print_columns(df) # 이미 위에서 보여줌
    
    cols = df.columns.tolist()
    
    x_str = config.get('x_vars_str', "")
    inp = input(f"- 독립변수(X) [이전: '{x_str}']: ").strip()
    x_str = inp if inp else x_str
    
    y_str = config.get('y_var_str', "")
    inp = input(f"- 종속변수(Y) [이전: '{y_str}']: ").strip()
    y_str = inp if inp else y_str
    
    return x_str, y_str

# -----------------------------------------------------------------------------
# Main Execution
# -----------------------------------------------------------------------------
def main():
    print("=== Data_Prepare Program Start ===")
    setup_environment()
    config = load_config()
    
    # 1. File Load
    target_file = select_file()
    if not target_file: return
    
    # 2. Scope - Start Row
    start_row = get_start_row(config)
    
    try:
        df_raw = pd.read_excel(target_file, header=start_row-1)
    except Exception as e:
        print(f"Error: {e}")
        return

    # 3. Scope - Drop Columns
    drop_str = get_drop_columns(df_raw, config)
    if drop_str:
        d_idxs = parse_column_selection(drop_str, df_raw.shape[1])
        df_raw.drop(columns=[df_raw.columns[i] for i in d_idxs], inplace=True)

    # 4. Filter by Class
    df_raw = filter_by_class(df_raw, config)

    # 4-2. Variables
    idx_str = get_index_selection(df_raw, config)
    idx_idxs = parse_column_selection(idx_str, df_raw.shape[1])
    
    x_str, y_str = get_var_selection(df_raw, config)
    x_idxs = parse_column_selection(x_str, df_raw.shape[1])
    y_idxs = parse_column_selection(y_str, df_raw.shape[1])
    
    # Exclude Index columns from X and Y if selected effectively
    # (Actually user might select them in range, we should remove them safely)
    x_idxs = [i for i in x_idxs if i not in idx_idxs]
    y_idxs = [i for i in y_idxs if i not in idx_idxs]
    
    idx_cols = [df_raw.columns[i] for i in idx_idxs]
    x_cols = [df_raw.columns[i] for i in x_idxs]
    y_cols = [df_raw.columns[i] for i in y_idxs]
    
    if idx_cols:
        print(f"[Info] Index 변수 제외됨: {idx_cols}")
    
    if not x_cols:
        print("독립변수가 선택되지 않아 Index를 제외한 전체 열을 대상으로 한다.")
        all_idxs = set(range(df_raw.shape[1])) - set(idx_idxs)
        x_cols = [df_raw.columns[i] for i in sorted(list(all_idxs))]

    # Save Config (Update existing)
    config.update({
        'start_row': start_row, 
        'drop_columns_str': drop_str, 
        'index_vars_str': idx_str,
        'x_vars_str': x_str, 
        'y_var_str': y_str,
        'x_names': x_cols,
        'y_names': y_cols,
        'index_names': idx_cols
    })
    save_config(config)

    # 5. Imputation Method
    print("\n[Step 5] 결측치 대체 기법 선택")
    for i, m in enumerate(IMPUTATION_METHODS):
        print(f"{i+1}. {m}")
    
    try:
        m_idx = int(input("번호 선택: ")) - 1
        method = IMPUTATION_METHODS[m_idx]
    except:
        method = "MICE"
        print("기본값 MICE를 사용한다")

    # 5. Run Imputation
    # run_imputation returns validation on x_cols only
    df_imputed_x = run_imputation(df_raw, method, x_cols, y_cols)
    
    # Merge Result (Original Y + Imputed X + Index)
    # df_imputed_x might have only X cols updated, but structure depends on implementation.
    # run_imputation returns 'imputed_df' which starts as a copy of 'target_df' (only X cols).
    # Wait, look at run_imputation:
    # target_df = df[x_cols].copy() -> imputed_df is copy of target_df.
    # So it ONLY contains X Columns.
    
    df_final = df_raw.copy()
    # Update only X Columns
    df_final[x_cols] = df_imputed_x[x_cols]
    
    # 6. Save Results
    base_name = os.path.basename(target_file).split('.')[0]
    res_enm = os.path.join(RESULT_DIR, f"{base_name}_{method}_Imputed.xlsx")
    pdf_enm = os.path.join(RESULT_DIR, f"{base_name}_{method}_Report.pdf")
    
    # Save Highlighted Excel
    save_excel_highlighted(df_raw, df_final, res_enm)
    
    # 7. Generate Report (Filtered by X Columns)
    create_pdf_report(pdf_enm, df_raw, df_final, method, x_cols)

if __name__ == "__main__":
    main()
